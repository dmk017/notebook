import io
import os
import sys
from typing import Optional

from returns.result import Failure, Result, Success
from openpyxl import Workbook

from src.utils.io.types import FileBufferContent
from src.utils.list.flatten import flatten
from src.models.models_router_api_schema import ModelsListFilter
from src.config import get_settings
from src.objects.data.constants import (
    MULTIPLE_SYMBOL,
    REQUIRED_SYMBOL
)
from src.errors.errors import Errors
from src.models.data import models_repository
from src.models.data.models_schema import (ModelPropertyPayload,
                                                  ModelSchema)
from src.properties import properties_service
from src.properties.data.properties_schema import (PropertyPayload,
                                                            PropertySchema)
from src.auth.data.auth_repository import get_auth_user_by_id
from src.keycloak.keycloak_models import UserRole
from src.types.primitive_type_enum import PrimitiveTypeEnum
from src.utils.csv_creater import create_sample_csv


def add_model(name: str,
              properties: list[ModelPropertyPayload],
              user_id: str) -> Result[ModelSchema,
                                      Errors.MODEL_DUPLICATE_NAME_4001]:
    previous_version = models_repository.get_actual_model_by_name(name)
    if previous_version:
        return Failure(Errors.MODEL_DUPLICATE_NAME_4001)
    return Success(models_repository.add_model(name=name, properties=properties, user_id=user_id))


def update_model(name: str,
                 properties: list[ModelPropertyPayload],
                 user_id: str) -> Result[ModelSchema,
                                         Errors.MODEL_NOT_FOUND_PREV_4002]:
    previous_version = models_repository.get_actual_model_by_name(name)
    if not previous_version:
        return Failure(Errors.MODEL_NOT_FOUND_PREV_4002)
    previous_version_id = str(previous_version.id)
    actual_model = models_repository.add_model(name=name, properties=properties, user_id=user_id)
    return Success(models_repository.update_model(previous_version_id, {'next_id': str(actual_model.id)}))


def get_available_models_by_user_id(user_id: str, user_role: UserRole):
    if user_role == UserRole.ADMIN:
        return models_repository.get_models(page=1, limit=sys.maxsize, filters=ModelsListFilter(is_actual=False, is_deleted=None))
    access_models_names = get_auth_user_by_id(user_id).access_model_names
    return flatten(list(filter(
        lambda schema: schema is not None,
        map(
            models_repository.get_models_by_name,
            access_models_names
        )
    )))


def get_model_by_id(id: str) -> Optional[ModelSchema]:
    return models_repository.get_model_by_id(id=id)


def get_model_history(id: str) -> list[ModelSchema]:
    model_data = models_repository.get_model_by_id(id)
    return models_repository.get_models_by_name(model_data.name)


def get_models_by_property_ids(property_ids: list[str]) -> list[ModelSchema]:
    return models_repository.get_models_by_property_ids(property_ids=property_ids)


def recove_delete_model(id: str, deleted: bool) -> Result[ModelSchema, Errors.MODEL_NOT_FOUND_4003]:
    models = get_model_history(id)
    for model in models:
      result = models_repository.update_model(model.id, {"deleted": deleted})
    result = get_model_by_id(id)
    if result is None:
        return Failure(Errors.MODEL_NOT_FOUND_4003)
    return Success(result)


def create_model_template_file(id: str) -> str:
    model = models_repository.get_model_by_id(id)
    properties = list(
        map(
            lambda p: dict(
                data=properties_service.get_property_by_id(p.id),
                is_required=p.is_required
            ),
            model.properties
        )
    )
    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    work_sheet = workbook.create_sheet(model.name)
    start_col_num = 0
    for property in properties:
        primitive_field_names = [
            _build_header_cell_template_file(
                property=property,
                payload_property=primitive,
                is_required_model_property=property['is_required']
            ) for primitive in property['data'].properties
        ]
        for col_num, values in enumerate(primitive_field_names, start=1):
            for row_num, value in enumerate(values, start=1):
                work_sheet.cell(
                    row=row_num,
                    column=start_col_num + col_num,
                    value=value
                )
        start_col_num += len(primitive_field_names)

    buffer = io.BytesIO()
    workbook.save(buffer)

    return FileBufferContent(
        name=f"template-{str(model.name)}-{str(model.created_at)}.xlsx",
        buffer=buffer
    )


def _build_header_cell_template_file(
    property,
    is_required_model_property: bool,
    payload_property: PropertyPayload,
) -> list[str]:
    # header_delimeter = get_settings().file_templates_header_delimeter
    return [
        f"{property['data'].name}{REQUIRED_SYMBOL if property['is_required'] else ''}",
        f"{payload_property.name}{MULTIPLE_SYMBOL if payload_property.is_multiple else ''}{REQUIRED_SYMBOL if is_required_model_property and payload_property.is_required else ''}",
        f"Тип значаний: {_decode_header_primitive_type(payload_property.primitive_type)}"
    ]


def _decode_header_primitive_type(primitive_type: PrimitiveTypeEnum) -> Optional[str]:
    if primitive_type == PrimitiveTypeEnum.BOOL.value:
        return 'флаг (доступные значния: True или False)'
    if primitive_type == PrimitiveTypeEnum.DATE.value:
        return 'дата (пример: 12.07.2023)'
    if primitive_type == PrimitiveTypeEnum.STR.value:
        return 'строка (например: мама мыла раму)'
    if primitive_type == PrimitiveTypeEnum.NUMBER.value:
        return 'число (например: 123)'
    if primitive_type == PrimitiveTypeEnum.FILE.value:
        return 'id файла из хранилища (например: 64ae632173a30b73fdbce55c)'
    return None
