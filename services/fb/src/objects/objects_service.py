import io
import sys
import os
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook

from returns.pipeline import is_successful
from returns.result import Failure, Result, Success

from src.objects.data.constants import MULTIPLE_DELIMETER
from src.utils.io.types import FileBufferContent
from src.errors.errors import Errors
from src.models import models_service
from src.models.data.models_schema import ModelSchema
from src.auth.data import auth_repository
from src.objects.conversion_types_map.conversion_types_map import \
    conversion_type
from src.objects.data import objects_repository
from src.objects.data.objects_schema import (
    ObjectPropertyPayload,
    ObjectSchema,
    ObjectStatus,
    TPayloadData
)
from src.objects.objects_service_schema import PropertyPayloadValues
from src.objects.objects_router_api_schema import SearchFiltersList
from src.objects.utils.objects_build_filters import parse_filters
from src.properties import properties_service
from src.types.response import ReponseError
from src.utils.mongo import PydanticObjectId
from src.utils.types.dict import filter_nullable_values_from_dict
from src.models.data.models_repository import get_model_by_id
from src.properties.properties_service import (
  get_actual_properties_schema_by_name
)


def get_objects_by_filter(
    page: int,
    limit: int,
    filters: list[SearchFiltersList] = [],
    model_ids: list[str] = [],
    text: str = ''
):
    search_filters = parse_filters(filters)
    return objects_repository.get_objects(
        page,
        limit,
        filter_nullable_values_from_dict(search_filters['filtersOr']),
        filter_nullable_values_from_dict(search_filters['filtersAnd']),
        model_ids=list(map(str, model_ids)),
        text=text
    )


def get_objects_count_by_filter(
    filters: list[SearchFiltersList] = [],
    model_ids: list[str] = [],
    text: str = ''
):
    search_filters = parse_filters(filters)
    return objects_repository.get_objects_count(
        filter_nullable_values_from_dict(search_filters['filtersOr']),
        filter_nullable_values_from_dict(search_filters['filtersAnd']),
        model_ids=list(map(str, model_ids)),
        text=text
    )


def get_objects_ids_by_filters(filters: list[SearchFiltersList]):
    objects = get_objects_by_filter(
        page=1,
        limit=sys.maxsize,
        filters=filters
    )
    return list(map(lambda item: str(item.id), objects))


def change_status_objects(action: str, ids: list[str], user_id: str, reason=''):
    return objects_repository.update_many_objects(
        list(map(PydanticObjectId, ids)),
        ObjectStatus(
            approve_at=datetime.utcnow() if action == 'approve' else None,
            decline_at=datetime.utcnow() if action == 'decline' else None,
            moderator_id=user_id,
            reason=reason
        )
    )


def add_object(
        owner_id: str,
        model_id: str,
        properties: list[PropertyPayloadValues]
) -> Result[ObjectSchema, list[ReponseError]]:
    validate = validate_object(
        owner_id=owner_id,
        model_id=model_id,
        properties=properties
    )

    if not is_successful(validate):
        return Failure(validate.failure())

    new_object = objects_repository.add_object(
        owner_id=owner_id,
        model_id=model_id,
        payload=validate.unwrap()
    )
    return Success(new_object)


def validate_object(
        owner_id: str,
        model_id: str,
        properties: list[PropertyPayloadValues]
) -> Result[list[ObjectPropertyPayload], list[ReponseError]]:
    payloads = []
    errors = []

    owner = auth_repository.get_auth_user_by_id(owner_id)
    if owner is None:
        errors.append(
            ReponseError(
                code=Errors.USER_NOT_FOUND_1001
            )
        )

    model = models_service.get_model_by_id(model_id)
    if model is None:
        errors.append(
            ReponseError(
                code=Errors.MODEL_NOT_FOUND_4003
            )
        )

    if not properties:
        errors.append(
            ReponseError(
                code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5005
            )
        )

    for property in properties:
        result = _encode_object_property(model, property)
        if is_successful(result):
            payloads.append(result.unwrap())
        else:
            errors.append(result.failure())

    if len(errors):
        return Failure(errors)

    return Success(payloads)


def _encode_object_property(
        model: ModelSchema,
        value: PropertyPayloadValues
) -> Result[ObjectPropertyPayload, ReponseError]:
    property = properties_service.get_actual_properties_schema_by_name(value.property_name)
    if property is None:
        return Failure(ReponseError(
            code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5001.value,
            details=f"{value.property_name}"
        ))

    property_data = []
    for data in value.data:
        property_payload = None
        for property_value in property.properties:
            if property_value.name == data.name:
                property_payload = property_value

        if property_payload is None or property_payload.primitive_type is None:
            return Failure(ReponseError(
                code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5002.value,
                details=f"{property.name} -> {data.name}"
            ))

        if not data.values:
            return Failure(ReponseError(
                code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5003.value,
                details=f"Значения для {data.name} не могут быть пустыми."
            ))

        values = list(filter(lambda v: bool(v), data.values))

        is_required_model_property = False
        for model_property in model.properties:
            if str(model_property.id) == str(property.id):
                is_required_model_property = model_property.is_required
                break

        if is_required_model_property and property_payload.is_required and not values:
            return Failure(ReponseError(
                code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5003.value,
                details=f"{property.name} -> {data.name}"
            ))

        if not property_payload.is_multiple and len(values) > 1:
            return Failure(ReponseError(
                code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5004.value,
                details=f"{property.name} -> {data.name}"
            ))

        payload_values = []
        pattern = re.compile(property_payload.validation)
        for v in values:
            if not pattern.match(str(v)):
                return Failure(ReponseError(
                    code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5007.value,
                    details=f"{property.name} -> {data.name}: {value} does not match validation pattern {property_payload.validation}"
                ))
            result = conversion_type(property_payload.primitive_type, v)
            if result is None:
                return Failure(ReponseError(
                    code=Errors.OBJECT_ADD_REQUEST_MALFORMED_5006.value,
                    details=f"{property.name} -> {data.name}: {v} is not {property_payload.primitive_type}"
                ))
            payload_values.append(result)

        item = TPayloadData.parse_obj({
            **data.dict(),
            'type': property_payload.primitive_type,
            'values': payload_values
        })
        property_data.append(item)
    return Success(ObjectPropertyPayload(
        property_name=value.property_name,
        data=property_data
    ))


def create_file_unloading(model_id: str, owner_name: str, objects: list[ObjectSchema]) -> FileBufferContent:
    model = get_model_by_id(model_id)
    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    sheets = {}

    for object in objects:
        for property in object.payload:
            if property.property_name not in sheets.keys():
                work_sheet = workbook.create_sheet(property.property_name)
                property_schema = get_actual_properties_schema_by_name(property.property_name)
                field_names = [property.name for property in property_schema.properties]
                for col_num, header in enumerate(field_names, start=1):
                    work_sheet.cell(row=1, column=col_num, value=header)
                sheets[property.property_name] = {'sheet': work_sheet, 'count_row': 1, 'field_names': field_names}

            sheets[property.property_name]['count_row'] += 1
            for col_num, header in enumerate(sheets[property.property_name]['field_names'], start=1):
                for filed in property.data:
                    if filed.name == header:
                        sheets[property.property_name]['sheet'].cell(row=sheets[property.property_name]['count_row'], column=col_num, value=MULTIPLE_DELIMETER.join(filed.values))
                        break

    buffer = io.BytesIO()
    workbook.save(buffer)

    return FileBufferContent(
        name=f"{owner_name}-{str(model.name)}-{str(model.created_at)}.xlsx",
        buffer=buffer
    )


def grouping_objects_by_attribute(attribute: str, objects: list[ObjectSchema]) -> defaultdict:
    grouped_objects = defaultdict(list)
    for object in objects:
        grouped_objects[str(dict(object)[attribute])].append(object)
    return grouped_objects


def cleanup_files(file_paths):
    for file_path in file_paths:
        os.remove(file_path)
